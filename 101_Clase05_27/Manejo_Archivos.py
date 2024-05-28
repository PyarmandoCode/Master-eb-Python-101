import csv
#pip install openpyxl
#import openpyxl 
from openpyxl import load_workbook

def leer_txt():
    try:
        archivo=open("Zenpython.txt","r",encoding='utf-8')
        contenido=archivo.read()
        print(contenido)
    except FileNotFoundError:
        print("Error el Archivo No Existe")    
    finally:
        #archivo.close()
        print("Archivo Cerrado")    


def leer_linea_linea_txt():
    try:
        with open("Zenpython.txt","r",encoding='utf-8') as archivo:
            for linea in archivo:
                print(linea,end='') #el Argumento end evita agregar un salto de linea adicional
    except FileNotFoundError:
        print("Error el Archivo No Existe")    
    finally:
        archivo.close()
        print("Archivo Cerrado")              

def leer_csv():
    try:
        with open('articulos.csv',encoding='utf-8') as archivo_csv:
            lector_csv=csv.reader(archivo_csv)
            lista=list()
            for linea in lector_csv:
                lista.append(linea)
    except FileNotFoundError:
        print("Error El Archivo no existe")
    finally:
        archivo_csv.close()
        print(lista)
        print("Archivo Cerrado")    


def leer_xlsx():
    try:
        lib_empl=load_workbook("empleados.xlsx")
        hoja=lib_empl.active
        for fila in hoja.iter_rows(values_only=True):
            print(fila) 
    except FileNotFoundError:
        print("Error el Archivo No Existe")    
    finally:
        print("Archivo Cerrado")    


lineas =[
     'Primera Linea del Archivo \n',
     'Segunda Linea del Archivo \n',
     'Tercera Linea del Archivo \n',
    ]

def escribir_txt():
    try:
        with open('archivo.txt','w',encoding='utf-8') as archivo:
            #Escribir los datos en el archivo
            archivo.writelines(lineas)
            #archivo.write(f'Edad : {edad}')
            #archivo.write(f'Ciudad : {ciudad}')
    except Exception as err:
        print(f"Ocurrio un error {err}")        
    finally:
        archivo.close()
        print("Archivo Cerrado correctamente")    


escribir_txt()





